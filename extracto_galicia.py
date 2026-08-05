"""Lectura del extracto de cuenta que exporta el Home Banking del Banco Galicia.

Galicia no tiene API pública de movimientos, así que la puerta de entrada de
esta variante de la app es el Excel que la persona se baja del Home Banking
(Cuentas -> Movimientos -> Exportar). Este módulo lo traduce a la misma forma
de "candidato a facturar" que la app de Mercado Pago arma desde su API.

Formato del archivo (verificado contra un extracto real de Caja de Ahorro):

    Fila 1-5   metadatos sueltos en la columna A ("Banco Galicia - Caja Ahorro
               Pesos", "Nro. de Cuenta: ...", "Intervalo de Consulta: ...").
    Fila 6     encabezados: Fecha | Movimiento | Débito | Crédito | Saldo
               Parcial | Comentarios.
    Fila 7+    un movimiento por fila. "Movimiento" es un bloque de texto de
               varias líneas: la primera es el concepto ("TRANSFERENCIA DE
               TERCEROS"), la segunda el nombre del titular que transfirió, la
               tercera su CUIT/CUIL, y el resto son CBUs, números de operación
               y ruido ("VARIOS", máscaras de tarjeta).

Nada de esto está documentado por el banco: si en algún momento cambian el
formato, lo que se rompe es la detección de encabezados y ahí conviene mirar
`_ubicar_encabezados`.
"""
import datetime
import hashlib
import re
import unicodedata

import openpyxl

# Encabezado -> rol de la columna. Se comparan normalizados (sin acentos, en
# minúscula) porque el banco alterna "Débito"/"Debito" según la exportación.
COLUMNAS = {
    "fecha": "fecha",
    "movimiento": "movimiento",
    "debito": "debito",
    "credito": "credito",
    "saldo parcial": "saldo",
    "saldo": "saldo",
    "comentarios": "comentarios",
    "comentario": "comentarios",
}

# Filas a mirar buscando el encabezado antes de rendirse.
MAX_FILAS_ENCABEZADO = 30

# Líneas del bloque "Movimiento" que nunca son el nombre del titular.
RUIDO = {"varios", "cuenta propia", "haberes", "sin comentarios"}

# Prefijos de CUIT/CUIL de personas físicas: solo de estos se puede deducir el
# DNI (los 30/33/34 son empresas y no tienen DNI atrás).
PREFIJOS_PERSONA = {"20", "23", "24", "27"}


def _normalizar(texto):
    texto = unicodedata.normalize("NFKD", str(texto or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.lower().split())


def _parsear_importe(valor):
    """Convierte "77.900,00" (formato del banco) a float. Devuelve 0.0 si no hay importe."""
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace("$", "").replace(" ", "")
    negativo = texto.startswith("-")
    texto = texto.lstrip("-")
    # Formato argentino: el punto es separador de miles y la coma, decimal.
    texto = texto.replace(".", "").replace(",", ".")
    try:
        numero = float(texto)
    except ValueError:
        return 0.0
    return -numero if negativo else numero


def _parsear_fecha(valor):
    """Acepta la fecha como texto DD/MM/AAAA (lo que manda Galicia) o como fecha real de Excel."""
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor

    texto = str(valor or "").strip()
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def cuit_valido(cuit):
    """Verifica el dígito verificador (módulo 11). Evita mandar a ARCA un CUIT
    mal leído del extracto: ARCA lo rechazaría recién al pedir el CAE."""
    digitos = re.sub(r"\D", "", str(cuit or ""))
    if len(digitos) != 11:
        return False
    pesos = (5, 4, 3, 2, 7, 6, 5, 4, 3, 2)
    suma = sum(int(d) * p for d, p in zip(digitos[:10], pesos))
    resto = suma % 11
    verificador = {0: 0, 1: 9}.get(resto, 11 - resto)
    return verificador == int(digitos[10])


def dni_desde_cuil(cuit):
    """Saca el DNI de adentro del CUIL de una persona física (27-13242063-2 -> 13242063).

    Devuelve None para CUITs de empresa (30/33/34) o si el CUIT no es válido."""
    digitos = re.sub(r"\D", "", str(cuit or ""))
    if not cuit_valido(digitos) or digitos[:2] not in PREFIJOS_PERSONA:
        return None
    return int(digitos[2:10])


def formatear_cuit(cuit):
    digitos = re.sub(r"\D", "", str(cuit or ""))
    if len(digitos) != 11:
        return digitos
    return f"{digitos[:2]}-{digitos[2:10]}-{digitos[10]}"


def _parece_nombre(linea):
    """El bloque de movimiento mezcla el nombre del titular con CBUs, números de
    operación y máscaras de tarjeta. El nombre es la primera línea que tiene
    letras de verdad."""
    texto = linea.strip()
    if len(texto) < 3:
        return False
    if _normalizar(texto) in RUIDO:
        return False
    # Descarta CBUs, números de operación y máscaras tipo 4517XXXXXXXXXX54.
    if re.fullmatch(r"[\dXx\s/.\-]+", texto):
        return False
    return any(c.isalpha() for c in texto)


def _desglosar_movimiento(texto):
    """Parte el bloque de texto del banco en concepto + titular + CUIT."""
    lineas = [l.strip() for l in str(texto or "").splitlines()]
    lineas = [l for l in lineas if l]

    concepto = lineas[0] if lineas else ""
    resto = lineas[1:]

    cuit = ""
    for linea in resto:
        candidato = re.sub(r"[\s.\-]", "", linea)
        # Solo líneas que son exclusivamente el CUIT: los CBUs tienen 22
        # dígitos y los números de operación no llegan a 11.
        if len(candidato) == 11 and candidato.isdigit():
            cuit = candidato
            break

    titular = ""
    for linea in resto:
        if _parece_nombre(linea):
            titular = linea
            break

    return {
        "concepto": concepto,
        "titular": titular,
        "cuit": cuit,
        "detalle_banco": "\n".join(lineas),
    }


def _ubicar_encabezados(ws):
    """Busca la fila de encabezados y devuelve (fila, {rol: indice_columna})."""
    for fila_nro, fila in enumerate(ws.iter_rows(min_row=1, max_row=MAX_FILAS_ENCABEZADO, values_only=True), start=1):
        mapa = {}
        for indice, celda in enumerate(fila):
            rol = COLUMNAS.get(_normalizar(celda))
            if rol and rol not in mapa:
                mapa[rol] = indice
        if "fecha" in mapa and "movimiento" in mapa and "credito" in mapa:
            return fila_nro, mapa
    return None, {}


def _metadatos(ws, hasta_fila):
    """Lee los datos sueltos de arriba del extracto (cuenta, período consultado)."""
    datos = {"banco": "", "cuenta": "", "periodo": ""}
    for fila in ws.iter_rows(min_row=1, max_row=max(hasta_fila - 1, 1), max_col=1, values_only=True):
        texto = str(fila[0] or "").strip()
        if not texto:
            continue
        normalizado = _normalizar(texto)
        if normalizado.startswith("nro. de cuenta") or normalizado.startswith("nro de cuenta"):
            datos["cuenta"] = texto.split(":", 1)[-1].strip()
        elif normalizado.startswith("intervalo de consulta"):
            datos["periodo"] = texto.split(":", 1)[-1].strip()
        elif "banco" in normalizado and not datos["banco"]:
            datos["banco"] = texto
    return datos


def _id_movimiento(fecha, monto, detalle, ocurrencia):
    """ID estable para un movimiento, para poder recordar cuáles ya se facturaron.

    El extracto no trae un número de operación propio, así que se arma un hash
    con lo que sí identifica a la fila (fecha + importe + bloque de texto). El
    `ocurrencia` distingue dos movimientos idénticos el mismo día — pasa cuando
    alguien transfiere dos veces lo mismo. A propósito no entra el saldo
    parcial: se quiere que el mismo movimiento tenga el mismo ID aunque el
    extracto se exporte con otro rango de fechas.
    """
    base = f"{fecha.isoformat()}|{monto:.2f}|{' '.join(detalle.split())}|{ocurrencia}"
    return "gal-" + hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def _parsear_hoja(ws):
    fila_encabezado, columnas = _ubicar_encabezados(ws)
    if not fila_encabezado:
        return None, [], 0

    cuenta = _metadatos(ws, fila_encabezado)
    cuenta["hoja"] = ws.title

    movimientos = []
    debitos = 0
    vistos = {}

    for fila in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
        def celda(rol):
            indice = columnas.get(rol)
            if indice is None or indice >= len(fila):
                return None
            return fila[indice]

        fecha = _parsear_fecha(celda("fecha"))
        if fecha is None:
            continue

        credito = _parsear_importe(celda("credito"))
        debito = _parsear_importe(celda("debito"))
        if credito <= 0:
            if debito > 0:
                debitos += 1
            continue

        desglose = _desglosar_movimiento(celda("movimiento"))
        clave = (fecha, round(credito, 2), " ".join(desglose["detalle_banco"].split()))
        vistos[clave] = vistos.get(clave, 0) + 1

        cuit = desglose["cuit"] if cuit_valido(desglose["cuit"]) else ""
        comentario = str(celda("comentarios") or "").strip()

        movimientos.append({
            "id": _id_movimiento(fecha, credito, desglose["detalle_banco"], vistos[clave]),
            "fecha": fecha.isoformat(),
            "monto": round(credito, 2),
            "de": desglose["titular"] or "(sin nombre en el extracto)",
            "cuit": cuit,
            "cuit_formateado": formatear_cuit(cuit),
            "dni": dni_desde_cuil(cuit),
            "concepto_banco": desglose["concepto"],
            "es_transferencia": "transferen" in _normalizar(desglose["concepto"]),
            "comentario": comentario,
            "detalle_banco": desglose["detalle_banco"],
            "saldo": _parsear_importe(celda("saldo")),
            "cuenta": cuenta.get("cuenta", ""),
        })

    return cuenta, movimientos, debitos


def leer_extracto(ruta):
    """Lee el Excel del Home Banking y devuelve los créditos recibidos.

    Devuelve {"cuentas": [...], "movimientos": [...], "debitos_ignorados": N}.
    Los movimientos vienen ordenados de más nuevo a más viejo y solo incluyen
    créditos (plata que entró): un débito nunca se factura.
    """
    if str(ruta).lower().endswith(".xls"):
        raise ValueError(
            "El archivo está en el formato viejo .xls. Abrilo con Excel y guardalo "
            "como .xlsx (Libro de Excel), o volvé a exportarlo desde el Home Banking "
            "eligiendo Excel."
        )

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(
            "No se pudo abrir el archivo como Excel. Asegurate de elegir el archivo "
            f"tal como lo baja el Home Banking de Galicia ({e})."
        ) from e

    try:
        cuentas = []
        movimientos = []
        debitos = 0
        for ws in wb.worksheets:
            cuenta, movs, debs = _parsear_hoja(ws)
            if cuenta is None:
                continue
            cuentas.append(cuenta)
            movimientos.extend(movs)
            debitos += debs
    finally:
        wb.close()

    if not cuentas:
        raise ValueError(
            "No se encontró la tabla de movimientos en el archivo. Tiene que ser el "
            "Excel de movimientos que exporta el Home Banking de Galicia (con las "
            "columnas Fecha, Movimiento, Débito y Crédito)."
        )

    movimientos.sort(key=lambda m: (m["fecha"], m["monto"]), reverse=True)
    return {"cuentas": cuentas, "movimientos": movimientos, "debitos_ignorados": debitos}


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Uso: python extracto_galicia.py <extracto.xlsx>")
        raise SystemExit(1)

    datos = leer_extracto(sys.argv[1])
    for cuenta in datos["cuentas"]:
        print(f"{cuenta['banco']} — cuenta {cuenta['cuenta']} — {cuenta['periodo']}")
    print(f"\n{len(datos['movimientos'])} créditos ({datos['debitos_ignorados']} débitos ignorados):\n")
    for m in datos["movimientos"]:
        marca = "" if m["es_transferencia"] else "  [no es transferencia]"
        print(f"  {m['fecha']}  ${m['monto']:>12,.2f}  {m['de'][:32]:<32} {m['cuit_formateado']:<14}{marca}")
